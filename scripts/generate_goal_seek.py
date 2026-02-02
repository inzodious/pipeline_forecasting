"""
Goal-Seek Baseline Exporter

Exports all historical metrics needed for Excel-based goal-seek forecasting.
Produces a formatted workbook ready for C-Suite presentation and scenario modeling.

Sheets:
1. Summary - One row per segment with all baseline assumptions
2. T12_Volume - Monthly volume by segment (trailing 12 months)
3. Win_Rates - Revenue-based win rates by segment
4. Stage_Probabilities - P(Win) by segment/stage with sample sizes
5. Open_Pipeline - Deal-level data for active pipeline going into 2026
6. Timing_Distribution - Months-to-close % by segment
7. Goal_Seek - Input cells for targets with formula-driven forecast
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os

# =============================================================================
# CONFIG
# =============================================================================

DATA_PATH = './data/fact_snapshots.csv'
OUTPUT_PATH = './exports/goal_seek_baseline.xlsx'
ACTUALS_THROUGH = '2025-12-26'
SKIPPER_WEIGHT = 0.5
CREDIBILITY_THRESHOLD = 50

# Styling
BLUE_FONT = Font(color='0000FF', bold=False)
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
INPUT_FILL = PatternFill('solid', fgColor='FFFF00')
CURRENCY_FORMAT = '$#,##0'
CURRENCY_FORMAT_M = '$#,##0.00,,"M"'
PERCENT_FORMAT = '0.0%'
NUMBER_FORMAT = '#,##0.0'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# DATA LOADING
# =============================================================================

def load_snapshots():
    df = pd.read_csv(DATA_PATH, parse_dates=['date_created', 'date_closed', 'date_snapshot'])
    df = df.sort_values(['deal_id', 'date_snapshot'])
    df['is_closed_won'] = df['stage'].isin(['Closed Won', 'Verbal'])
    df['is_closed_lost'] = df['stage'] == 'Closed Lost'
    df['is_closed'] = df['is_closed_won'] | df['is_closed_lost']
    return df

def build_deal_facts(df):
    first = df.groupby('deal_id').first().reset_index()
    closed = df[df['is_closed']].groupby('deal_id').first().reset_index()[['deal_id', 'date_closed', 'stage', 'net_revenue']]
    closed = closed.rename(columns={'stage': 'final_stage', 'date_closed': 'final_date_closed', 'net_revenue': 'final_net_revenue'})
    
    deals = first.merge(closed, on='deal_id', how='left')
    deals['created_month'] = deals['date_created'].dt.to_period('M')
    deals['stage'] = deals['final_stage'].fillna('Open')
    deals['date_closed'] = deals['final_date_closed']
    deals['net_revenue'] = deals['final_net_revenue'].fillna(deals['net_revenue'])
    deals['won'] = deals['stage'].isin(['Closed Won', 'Verbal'])
    
    first_stage = df.groupby('deal_id')['stage'].first()
    skippers = first_stage[first_stage == 'Closed Lost'].index
    deals['volume_weight'] = 1.0
    deals.loc[deals['deal_id'].isin(skippers), 'volume_weight'] = SKIPPER_WEIGHT
    
    return deals

# =============================================================================
# METRIC CALCULATIONS
# =============================================================================

def calc_t12_volume(deals):
    cutoff = deals['date_created'].max()
    t12_start = cutoff - pd.DateOffset(months=12)
    t12_deals = deals[deals['date_created'] >= t12_start]
    
    monthly = t12_deals.groupby(['market_segment', 'created_month']).agg(
        raw_count=('deal_id', 'count'),
        adj_volume=('volume_weight', 'sum'),
        total_revenue=('net_revenue', 'sum')
    ).reset_index()
    monthly['month'] = monthly['created_month'].astype(str)
    
    summary = monthly.groupby('market_segment').agg(
        months_in_t12=('month', 'count'),
        total_raw_count=('raw_count', 'sum'),
        total_adj_volume=('adj_volume', 'sum'),
        total_revenue=('total_revenue', 'sum')
    ).reset_index()
    summary['avg_monthly_volume'] = summary['total_adj_volume'] / summary['months_in_t12']
    summary['avg_monthly_revenue'] = summary['total_revenue'] / summary['months_in_t12']
    
    return monthly[['month', 'market_segment', 'raw_count', 'adj_volume', 'total_revenue']], summary

def calc_win_rates(deals):
    cutoff = deals['date_created'].max()
    t12_start = cutoff - pd.DateOffset(months=12)
    t12_deals = deals[deals['date_created'] >= t12_start]
    
    def segment_metrics(df):
        total_rev = (df['net_revenue'] * df['volume_weight']).sum()
        won_rev = (df[df['won']]['net_revenue'] * df[df['won']]['volume_weight']).sum()
        win_rate = won_rev / total_rev if total_rev > 0 else 0
        
        total_count = df['volume_weight'].sum()
        won_count = df[df['won']]['volume_weight'].sum()
        win_rate_count = won_count / total_count if total_count > 0 else 0
        
        avg_deal_size = df[df['won']]['net_revenue'].mean() if df['won'].sum() > 0 else 0
        
        return pd.Series({
            'total_pipeline_revenue': total_rev,
            'won_revenue': won_rev,
            'win_rate_revenue': win_rate,
            'total_deals': total_count,
            'won_deals': won_count,
            'win_rate_count': win_rate_count,
            'avg_won_deal_size': avg_deal_size,
            'sample_size': len(df)
        })
    
    t12_wr = t12_deals.groupby('market_segment').apply(segment_metrics).reset_index()
    t12_wr.columns = ['market_segment'] + ['t12_' + c for c in t12_wr.columns[1:]]
    
    all_wr = deals.groupby('market_segment').apply(segment_metrics).reset_index()
    all_wr.columns = ['market_segment'] + ['all_' + c for c in all_wr.columns[1:]]
    
    merged = t12_wr.merge(all_wr, on='market_segment')
    merged['blended_win_rate'] = (merged['t12_win_rate_revenue'] * 0.55) + (merged['all_win_rate_revenue'] * 0.45)
    merged['blended_deal_size'] = (merged['t12_avg_won_deal_size'] * 0.55) + (merged['all_avg_won_deal_size'] * 0.45)
    
    return merged

def calc_stage_probabilities(df):
    df = df.sort_values(['deal_id', 'date_snapshot'])
    first_closed = df[df['is_closed']].groupby('deal_id').first().reset_index()
    first_closed = first_closed[['deal_id', 'date_snapshot', 'market_segment', 'is_closed_won', 'stage']]
    first_closed = first_closed.rename(columns={'date_snapshot': 'close_snap', 'stage': 'close_stage'})
    
    prev_snap = df.merge(first_closed[['deal_id', 'close_snap']], on='deal_id', how='inner')
    prev_snap = prev_snap[prev_snap['date_snapshot'] < prev_snap['close_snap']]
    stage_before_exit = prev_snap.groupby('deal_id').last().reset_index()[['deal_id', 'stage']]
    stage_before_exit = stage_before_exit.rename(columns={'stage': 'stage_at_exit'})
    
    exits = first_closed.merge(stage_before_exit, on='deal_id', how='left')
    exits['stage_at_exit'] = exits['stage_at_exit'].fillna(exits['close_stage'])
    exits['stage'] = exits['stage_at_exit']
    
    cutoff_date = df['date_snapshot'].max()
    t12_start = cutoff_date - pd.DateOffset(months=12)
    exits['weight'] = np.where(exits['close_snap'] >= t12_start, 3.0, 1.0)
    
    probs_list = []
    for (segment, stage), group in exits.groupby(['market_segment', 'stage']):
        wins = (group['is_closed_won'] * group['weight']).sum()
        total = group['weight'].sum()
        unweighted_total = len(group)
        prob = wins / total if total > 0 else 0
        probs_list.append({
            'market_segment': segment,
            'stage': stage,
            'weighted_wins': wins,
            'weighted_total': total,
            'unweighted_exits': unweighted_total,
            'raw_probability': prob
        })
    
    probs = pd.DataFrame(probs_list)
    
    global_list = []
    for stage, group in exits.groupby('stage'):
        g_wins = (group['is_closed_won'] * group['weight']).sum()
        g_total = group['weight'].sum()
        global_list.append({'stage': stage, 'global_prob': g_wins / g_total if g_total > 0 else 0})
    global_probs = pd.DataFrame(global_list)
    
    probs = probs.merge(global_probs, on='stage', how='left')
    probs['credibility'] = (probs['weighted_total'] / CREDIBILITY_THRESHOLD).clip(0, 1.0)
    probs['blended_probability'] = (probs['raw_probability'] * probs['credibility']) + (probs['global_prob'].fillna(0) * (1 - probs['credibility']))
    probs['final_probability'] = np.maximum(probs['blended_probability'], probs['global_prob'].fillna(0) * 0.95)
    
    return probs

def calc_open_pipeline(df, stage_probs):
    cutoff_dt = pd.to_datetime(ACTUALS_THROUGH)
    available = df[df['date_snapshot'] <= cutoff_dt]
    last_date = available['date_snapshot'].max()
    
    open_deals = df[(df['date_snapshot'] == last_date) & (~df['is_closed'])].copy()
    open_deals['age_weeks'] = ((open_deals['date_snapshot'] - open_deals['date_created']).dt.days // 7)
    open_deals['age_days'] = (open_deals['date_snapshot'] - open_deals['date_created']).dt.days
    
    open_deals = open_deals.merge(
        stage_probs[['market_segment', 'stage', 'final_probability']],
        on=['market_segment', 'stage'],
        how='left'
    )
    open_deals['probability'] = open_deals['final_probability'].fillna(0)
    open_deals['expected_revenue'] = open_deals['net_revenue'] * open_deals['probability']
    
    return open_deals[['deal_id', 'market_segment', 'stage', 'net_revenue', 'age_days', 'age_weeks', 'probability', 'expected_revenue', 'date_created']]

def calc_timing_distribution(deals):
    won_deals = deals[deals['won'] & deals['date_closed'].notna()].copy()
    won_deals['months_to_close'] = ((won_deals['date_closed'] - won_deals['date_created']).dt.days / 30).clip(0, 11).round().astype(int)
    
    dist = won_deals.groupby(['market_segment', 'months_to_close']).size().reset_index(name='count')
    totals = dist.groupby('market_segment')['count'].sum().reset_index(name='total')
    dist = dist.merge(totals, on='market_segment')
    dist['pct'] = dist['count'] / dist['total']
    
    return dist

def build_summary(deals, win_rates, t12_summary, open_pipeline):
    summary = t12_summary[['market_segment', 'avg_monthly_volume']].copy()
    summary = summary.merge(
        win_rates[['market_segment', 'blended_win_rate', 'blended_deal_size', 't12_sample_size']],
        on='market_segment'
    )
    
    pipeline_summary = open_pipeline.groupby('market_segment').agg(
        open_deals=('deal_id', 'count'),
        open_pipeline_value=('net_revenue', 'sum'),
        expected_pipeline_value=('expected_revenue', 'sum')
    ).reset_index()
    
    summary = summary.merge(pipeline_summary, on='market_segment', how='left')
    
    summary['baseline_annual_revenue'] = (
        summary['avg_monthly_volume'] * 12 * 
        summary['blended_win_rate'] * 
        summary['blended_deal_size']
    )
    
    return summary

# =============================================================================
# EXCEL FORMATTING HELPERS
# =============================================================================

def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def auto_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = None
        for cell in column_cells:
            if hasattr(cell, 'column_letter'):
                column = cell.column_letter
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        if column:
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column].width = adjusted_width

def add_dataframe_to_sheet(ws, df, start_row=1, currency_cols=None, pct_cols=None, number_cols=None):
    currency_cols = currency_cols or []
    pct_cols = pct_cols or []
    number_cols = number_cols or []
    
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col_name)
    style_header_row(ws, start_row, len(df.columns))
    
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            
            col_name = df.columns[c_idx - 1]
            if col_name in currency_cols:
                cell.number_format = CURRENCY_FORMAT
            elif col_name in pct_cols:
                cell.number_format = PERCENT_FORMAT
            elif col_name in number_cols:
                cell.number_format = NUMBER_FORMAT

# =============================================================================
# BUILD GOAL SEEK SHEET
# =============================================================================

def build_goal_seek_sheet(ws, summary, segments):
    ws['A1'] = 'FY2026 Goal-Seek Forecast Model'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # Section: Baseline Metrics (from data)
    ws['A3'] = 'BASELINE METRICS (From Historical Data)'
    ws['A3'].font = Font(bold=True, size=12)
    
    headers = ['Segment', 'T12 Avg Monthly Vol', 'Blended Win Rate', 'Blended Deal Size', 
               'Open Pipeline ($)', 'Expected Pipeline ($)', 'Baseline Annual Revenue']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))
    
    seg_rows = {}
    for r, seg in enumerate(segments, 5):
        seg_data = summary[summary['market_segment'] == seg]
        if seg_data.empty:
            continue
        seg_data = seg_data.iloc[0]
        seg_rows[seg] = r
        
        ws.cell(row=r, column=1, value=seg)
        ws.cell(row=r, column=2, value=seg_data['avg_monthly_volume']).number_format = NUMBER_FORMAT
        ws.cell(row=r, column=3, value=seg_data['blended_win_rate']).number_format = PERCENT_FORMAT
        ws.cell(row=r, column=4, value=seg_data['blended_deal_size']).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=5, value=seg_data.get('open_pipeline_value', 0)).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=6, value=seg_data.get('expected_pipeline_value', 0)).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=7, value=seg_data['baseline_annual_revenue']).number_format = CURRENCY_FORMAT
        
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
    
    # Section: Scenario Levers (user inputs)
    lever_start = 5 + len(segments) + 2
    ws.cell(row=lever_start, column=1, value='SCENARIO LEVERS (Adjust for Goal Seek)')
    ws.cell(row=lever_start, column=1).font = Font(bold=True, size=12)
    
    lever_headers = ['Segment', 'Volume Mult', 'Win Rate Mult', 'Deal Size Mult', 'Revenue Target', 'Gap to Target']
    for c, h in enumerate(lever_headers, 1):
        ws.cell(row=lever_start + 1, column=c, value=h)
    style_header_row(ws, lever_start + 1, len(lever_headers))
    
    lever_rows = {}
    for i, seg in enumerate(segments):
        r = lever_start + 2 + i
        lever_rows[seg] = r
        
        ws.cell(row=r, column=1, value=seg)
        
        vol_cell = ws.cell(row=r, column=2, value=1.0)
        vol_cell.font = BLUE_FONT
        vol_cell.fill = INPUT_FILL
        vol_cell.number_format = '0.00'
        
        wr_cell = ws.cell(row=r, column=3, value=1.0)
        wr_cell.font = BLUE_FONT
        wr_cell.fill = INPUT_FILL
        wr_cell.number_format = '0.00'
        
        ds_cell = ws.cell(row=r, column=4, value=1.0)
        ds_cell.font = BLUE_FONT
        ds_cell.fill = INPUT_FILL
        ds_cell.number_format = '0.00'
        
        target_cell = ws.cell(row=r, column=5, value=0)
        target_cell.font = BLUE_FONT
        target_cell.fill = INPUT_FILL
        target_cell.number_format = CURRENCY_FORMAT
        
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
    
    # Section: Forecast Output (formulas)
    forecast_start = lever_start + 2 + len(segments) + 2
    ws.cell(row=forecast_start, column=1, value='FORECAST OUTPUT (Formula-Driven)')
    ws.cell(row=forecast_start, column=1).font = Font(bold=True, size=12)
    
    forecast_headers = ['Segment', 'Adj Monthly Vol', 'Adj Win Rate', 'Adj Deal Size', 
                        'Future Pipeline Rev', 'Active Pipeline Rev', 'Total Forecast']
    for c, h in enumerate(forecast_headers, 1):
        ws.cell(row=forecast_start + 1, column=c, value=h)
    style_header_row(ws, forecast_start + 1, len(forecast_headers))
    
    for i, seg in enumerate(segments):
        r = forecast_start + 2 + i
        base_r = seg_rows.get(seg, 5)
        lever_r = lever_rows.get(seg, lever_start + 2)
        
        ws.cell(row=r, column=1, value=seg)
        
        # Adj Monthly Vol = Baseline Vol * Vol Mult
        ws.cell(row=r, column=2, value=f'=B{base_r}*B{lever_r}').number_format = NUMBER_FORMAT
        
        # Adj Win Rate = Baseline WR * WR Mult (capped at 100%)
        ws.cell(row=r, column=3, value=f'=MIN(C{base_r}*C{lever_r}, 1)').number_format = PERCENT_FORMAT
        
        # Adj Deal Size = Baseline Size * Size Mult
        ws.cell(row=r, column=4, value=f'=D{base_r}*D{lever_r}').number_format = CURRENCY_FORMAT
        
        # Future Pipeline Rev = Adj Vol * 12 * Adj WR * Adj Size
        ws.cell(row=r, column=5, value=f'=B{r}*12*C{r}*D{r}').number_format = CURRENCY_FORMAT
        
        # Active Pipeline Rev (from baseline expected)
        ws.cell(row=r, column=6, value=f'=F{base_r}').number_format = CURRENCY_FORMAT
        
        # Total Forecast = Future + Active
        ws.cell(row=r, column=7, value=f'=E{r}+F{r}').number_format = CURRENCY_FORMAT
        
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
        
        # Update Gap formula in levers section
        gap_cell = ws.cell(row=lever_r, column=6, value=f'=E{lever_r}-G{r}')
        gap_cell.number_format = CURRENCY_FORMAT
    
    # Totals row
    total_r = forecast_start + 2 + len(segments)
    ws.cell(row=total_r, column=1, value='TOTAL').font = Font(bold=True)
    for c in range(5, 8):
        col_letter = get_column_letter(c)
        ws.cell(row=total_r, column=c, value=f'=SUM({col_letter}{forecast_start+2}:{col_letter}{total_r-1})')
        ws.cell(row=total_r, column=c).number_format = CURRENCY_FORMAT
        ws.cell(row=total_r, column=c).font = Font(bold=True)
    
    # Instructions
    instr_start = total_r + 3
    ws.cell(row=instr_start, column=1, value='HOW TO USE GOAL SEEK:')
    ws.cell(row=instr_start, column=1).font = Font(bold=True)
    ws.cell(row=instr_start + 1, column=1, value='1. Enter your revenue target in the yellow "Revenue Target" cell for a segment')
    ws.cell(row=instr_start + 2, column=1, value='2. Go to Data → What-If Analysis → Goal Seek')
    ws.cell(row=instr_start + 3, column=1, value='3. Set cell: The "Gap to Target" cell for that segment')
    ws.cell(row=instr_start + 4, column=1, value='4. To value: 0')
    ws.cell(row=instr_start + 5, column=1, value='5. By changing cell: One of the multiplier cells (Volume, Win Rate, or Deal Size)')
    ws.cell(row=instr_start + 6, column=1, value='6. Click OK - Excel will solve for the required multiplier')
    
    auto_width(ws)

# =============================================================================
# MAIN EXPORT
# =============================================================================

def main():
    print("Loading data...")
    snapshots = load_snapshots()
    deals = build_deal_facts(snapshots)
    
    print("Calculating T12 Volume...")
    t12_monthly, t12_summary = calc_t12_volume(deals)
    
    print("Calculating Win Rates...")
    win_rates = calc_win_rates(deals)
    
    print("Calculating Stage Probabilities...")
    stage_probs = calc_stage_probabilities(snapshots)
    
    print("Calculating Open Pipeline...")
    open_pipeline = calc_open_pipeline(snapshots, stage_probs)
    
    print("Calculating Timing Distribution...")
    timing_dist = calc_timing_distribution(deals)
    
    print("Building Summary...")
    summary = build_summary(deals, win_rates, t12_summary, open_pipeline)
    
    segments = sorted(deals['market_segment'].unique())
    
    print("Creating Excel workbook...")
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    add_dataframe_to_sheet(
        ws_summary, summary, 
        currency_cols=['blended_deal_size', 'open_pipeline_value', 'expected_pipeline_value', 'baseline_annual_revenue'],
        pct_cols=['blended_win_rate'],
        number_cols=['avg_monthly_volume', 't12_sample_size', 'open_deals']
    )
    auto_width(ws_summary)
    
    # Sheet 2: T12 Volume
    ws_vol = wb.create_sheet('T12_Volume')
    add_dataframe_to_sheet(
        ws_vol, t12_monthly,
        currency_cols=['total_revenue'],
        number_cols=['raw_count', 'adj_volume']
    )
    auto_width(ws_vol)
    
    # Sheet 3: Win Rates
    ws_wr = wb.create_sheet('Win_Rates')
    add_dataframe_to_sheet(
        ws_wr, win_rates,
        currency_cols=['t12_total_pipeline_revenue', 't12_won_revenue', 't12_avg_won_deal_size',
                       'all_total_pipeline_revenue', 'all_won_revenue', 'all_avg_won_deal_size', 'blended_deal_size'],
        pct_cols=['t12_win_rate_revenue', 't12_win_rate_count', 'all_win_rate_revenue', 'all_win_rate_count', 'blended_win_rate'],
        number_cols=['t12_total_deals', 't12_won_deals', 't12_sample_size', 'all_total_deals', 'all_won_deals', 'all_sample_size']
    )
    auto_width(ws_wr)
    
    # Sheet 4: Stage Probabilities
    ws_stage = wb.create_sheet('Stage_Probabilities')
    add_dataframe_to_sheet(
        ws_stage, stage_probs,
        pct_cols=['raw_probability', 'global_prob', 'credibility', 'blended_probability', 'final_probability'],
        number_cols=['weighted_wins', 'weighted_total', 'unweighted_exits']
    )
    auto_width(ws_stage)
    
    # Sheet 5: Open Pipeline
    ws_pipe = wb.create_sheet('Open_Pipeline')
    add_dataframe_to_sheet(
        ws_pipe, open_pipeline,
        currency_cols=['net_revenue', 'expected_revenue'],
        pct_cols=['probability'],
        number_cols=['age_days', 'age_weeks']
    )
    auto_width(ws_pipe)
    
    # Sheet 6: Timing Distribution
    ws_timing = wb.create_sheet('Timing_Distribution')
    add_dataframe_to_sheet(
        ws_timing, timing_dist,
        pct_cols=['pct'],
        number_cols=['months_to_close', 'count', 'total']
    )
    auto_width(ws_timing)
    
    # Sheet 7: Goal Seek
    ws_goal = wb.create_sheet('Goal_Seek')
    build_goal_seek_sheet(ws_goal, summary, segments)
    
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nExported to: {OUTPUT_PATH}")
    
    print("\n" + "="*60)
    print("SUMMARY OF EXPORTED METRICS")
    print("="*60)
    print(f"\nSegments: {', '.join(segments)}")
    print(f"Open Pipeline Deals: {len(open_pipeline)}")
    print(f"Open Pipeline Value: ${open_pipeline['net_revenue'].sum():,.0f}")
    print(f"Expected Pipeline Value: ${open_pipeline['expected_revenue'].sum():,.0f}")
    print(f"\nBaseline Annual Revenue by Segment:")
    for _, row in summary.iterrows():
        print(f"  {row['market_segment']}: ${row['baseline_annual_revenue']:,.0f}")
    print(f"\nTOTAL BASELINE: ${summary['baseline_annual_revenue'].sum():,.0f}")

if __name__ == '__main__':
    main()